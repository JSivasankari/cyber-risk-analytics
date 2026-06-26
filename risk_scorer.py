import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# R = w1*severity + w2*frequency + w3*scale + w4*lag_penalty
W_SEVERITY  = 0.35
W_FREQUENCY = 0.25
W_SCALE     = 0.25
W_LAG       = 0.15   

def normalize(series: pd.Series) -> pd.Series:
   
    lo, hi = series.min(), series.max()
    if hi == lo:
        return pd.Series([0.5] * len(series), index=series.index)
    return (series - lo) / (hi - lo)

def compute_industry_risk(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate breach data by industry and compute composite risk score.

    Risk Score (0–1):
        R = 0.35 * norm(avg_severity)
          + 0.25 * norm(breach_count)
          + 0.25 * norm(total_pwned_millions)
          + 0.15 * norm(avg_disclosure_lag)

    Higher lag → higher risk (company was slow to disclose).
    """
    grp = df.groupby("Industry").agg(
        BreachCount       = ("Name",               "count"),
        TotalPwnedM       = ("PwnMillions",         "sum"),
        AvgPwnedM         = ("PwnMillions",         "mean"),
        MaxPwnedM         = ("PwnMillions",         "max"),
        AvgSeverityScore  = ("SeverityScore",       "mean"),
        MaxSeverityScore  = ("SeverityScore",       "max"),
        CredentialLeaks   = ("IsCredentialLeak",    "sum"),
        AvgLagDays        = ("DisclosureLagDays",   "mean"),
        MaxLagDays        = ("DisclosureLagDays",   "max"),
        CriticalBreaches  = ("SeverityLabel",
                             lambda x: (x == "Critical").sum()),
        HighBreaches      = ("SeverityLabel",
                             lambda x: (x == "High").sum()),
        YearFirstSeen     = ("Year",                "min"),
        YearLastSeen      = ("Year",                "max"),
    ).reset_index()

    grp["SevNorm"]  = normalize(grp["AvgSeverityScore"])
    grp["FreqNorm"] = normalize(grp["BreachCount"])
    grp["ScaleNorm"]= normalize(grp["TotalPwnedM"])
    grp["LagNorm"]  = normalize(grp["AvgLagDays"].fillna(0))

    grp["RiskScore"] = (
        W_SEVERITY  * grp["SevNorm"]  +
        W_FREQUENCY * grp["FreqNorm"] +
        W_SCALE     * grp["ScaleNorm"]+
        W_LAG       * grp["LagNorm"]
    ).round(4)

    grp["RiskScorePct"] = (grp["RiskScore"] * 100).round(1)

    grp["RiskLabel"] = grp["RiskScore"].apply(
        lambda s: "Critical" if s > 0.70
             else "High"     if s > 0.45
             else "Medium"   if s > 0.25
             else "Low"
    )

    grp["CredLeakRate%"] = (
        grp["CredentialLeaks"] / grp["BreachCount"] * 100
    ).round(1)

  
    for col in ["TotalPwnedM", "AvgPwnedM", "MaxPwnedM",
                "AvgSeverityScore", "MaxSeverityScore",
                "AvgLagDays", "MaxLagDays"]:
        grp[col] = grp[col].round(2)

    grp = grp.drop(columns=["SevNorm", "FreqNorm", "ScaleNorm", "LagNorm"])

    return grp.sort_values("RiskScore", ascending=False).reset_index(drop=True)
def compute_yearly_trend(df: pd.DataFrame) -> pd.DataFrame:
    """Year-level aggregations for timeline visuals in Power BI."""
    trend = df.groupby("Year").agg(
        BreachCount      = ("Name",             "count"),
        TotalPwnedM      = ("PwnMillions",       "sum"),
        AvgSeverityScore = ("SeverityScore",     "mean"),
        CredLeaks        = ("IsCredentialLeak",  "sum"),
        UniqueIndustries = ("Industry",          "nunique"),
    ).reset_index()

    trend["CumulativePwnedM"] = trend["TotalPwnedM"].cumsum().round(2)
    trend["TotalPwnedM"]      = trend["TotalPwnedM"].round(2)
    trend["AvgSeverityScore"] = trend["AvgSeverityScore"].round(2)

    return trend.dropna(subset=["Year"]).sort_values("Year")


def compute_data_class_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten DataClassesStr into one row per (breach, data_class).
    Power BI can use this for the data-class frequency bar chart
    after an Unpivot step is no longer needed.
    """
    rows = []
    for _, row in df.iterrows():
        if not isinstance(row["DataClassesStr"], str):
            continue
        for cls in row["DataClassesStr"].split(","):
            cls = cls.strip()
            if cls:
                rows.append({
                    "BreachName":   row["Name"],
                    "Year":         row.get("Year"),
                    "Industry":     row.get("Industry"),
                    "SeverityLabel":row.get("SeverityLabel"),
                    "DataClass":    cls,
                })
    matrix = pd.DataFrame(rows)
    return matrix.groupby(["DataClass", "Industry", "SeverityLabel"]).agg(
        Frequency=("BreachName", "count")
    ).reset_index().sort_values("Frequency", ascending=False)


def style_excel(output_path: str):
    
    wb = load_workbook(output_path)

    HEADER_FILL  = PatternFill("solid", fgColor="1F2937")   # dark slate
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="F9FAFB")
    BORDER_SIDE  = Side(style="thin", color="E5E7EB")
    CELL_BORDER  = Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE,  bottom=BORDER_SIDE
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column
        max_row = ws.max_row

        
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = CELL_BORDER

        
        for row_idx in range(2, max_row + 1):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = CELL_BORDER
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="center")

       
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len    = 0
            for row_idx in range(1, min(max_row + 1, 200)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

       
        ws.freeze_panes = "A2"

        
        for col_idx in range(1, max_col + 1):
            header = ws.cell(row=1, column=col_idx).value or ""
            if "Risk" in str(header) or "Severity" in str(header):
                col_letter = get_column_letter(col_idx)
                col_range  = f"{col_letter}2:{col_letter}{max_row}"
                ws.conditional_formatting.add(
                    col_range,
                    ColorScaleRule(
                        start_type="min",  start_color="63BE7B",  # green
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max",    end_color="F8696B"   # red
                    )
                )

    wb.save(output_path)


def main():
    input_path  = os.path.join("data", "breaches_tagged.csv")
    output_path = os.path.join("data", "breaches_final.xlsx")

    print("[1/5] Loading tagged breach data...")
    df = pd.read_csv(input_path)
    print(f"      ✓ {len(df)} records loaded")

    print("[2/5] Computing industry risk scores...")
    risk_df = compute_industry_risk(df)
    print(f"      ✓ {len(risk_df)} industries scored")
    print(risk_df[["Industry", "RiskScore", "RiskLabel", "BreachCount"]].head(8).to_string(index=False))

    print("[3/5] Computing yearly trend data...")
    trend_df = compute_yearly_trend(df)

    print("[4/5] Computing data class frequency matrix...")
    matrix_df = compute_data_class_matrix(df)

    print(f"[5/5] Writing Power BI-ready Excel workbook → {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer,        sheet_name="Breaches",        index=False)
        risk_df.to_excel(writer,   sheet_name="IndustryRisk",    index=False)
        trend_df.to_excel(writer,  sheet_name="YearlyTrend",     index=False)
        matrix_df.to_excel(writer, sheet_name="DataClassMatrix", index=False)

    print("      Applying professional Excel formatting...")
    style_excel(output_path)

    print(f"""
✅ Done! breaches_final.xlsx exported with 4 sheets:
   • Breaches        — {len(df):,} breach records (full detail)
   • IndustryRisk    — {len(risk_df)} industries with composite risk scores
   • YearlyTrend     — {len(trend_df)} years of breach trend data
   • DataClassMatrix — {len(matrix_df):,} data class frequency rows


""")


if __name__ == "__main__":
    main()
